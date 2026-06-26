import os
from openpyxl import load_workbook


class ExcelReader:
    """Read test data from Excel files."""

    @staticmethod
    def read_data(file_name, sheet_name="Sheet1"):
        """
        Read all rows from an Excel sheet and return as list of dicts.

        Args:
            file_name: Name of Excel file (relative to TestData folder)
            sheet_name: Name of sheet to read (default: Sheet1)

        Returns:
            List of dictionaries where keys are column headers

        Example:
            data = ExcelReader.read_data("LoginData.xlsx")
            for row in data:
                username = row['username']
                password = row['password']
        """
        # Construct path to TestData folder
        test_data_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "TestData",
            file_name,
        )

        # Load workbook
        workbook = load_workbook(test_data_path)
        sheet = workbook[sheet_name]

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)

        # Read data rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if any(cell is not None for cell in row):
                row_dict = {}
                for i, header in enumerate(headers):
                    row_dict[header] = row[i]
                data.append(row_dict)

        workbook.close()
        return data
