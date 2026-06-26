# Script to create LoginData.xlsx with test credentials per market
from openpyxl import Workbook
import os

# Market test data
MARKET_DATA = {
    'de-de': [
        {'username': 'muhammed.najeeb+de_uat_adv@electrolux.com', 'password': 'Rebin@2002'},
    ],
    'nl-be': [
        {'username': 'testuser_nl@example.com', 'password': 'Pass@1234'},
    ],
    'fr-fr': [
        {'username': 'testuser_fr@example.com', 'password': 'Pass@5678'},
    ],
    'it-it': [
        {'username': 'testuser_it@example.com', 'password': 'Pass@9012'},
    ],
}

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Create a sheet for each market
for market, credentials in MARKET_DATA.items():
    ws = wb.create_sheet(title=market)

    # Add headers
    ws['A1'] = 'username'
    ws['B1'] = 'password'

    # Add test data for this market
    for i, cred in enumerate(credentials, start=2):
        ws[f'A{i}'] = cred['username']
        ws[f'B{i}'] = cred['password']

# Save file
test_data_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(test_data_dir, 'LoginData.xlsx')
wb.save(file_path)
wb.close()

print("✅ LoginData.xlsx created successfully!")
print(f"Location: {file_path}")
print(f"Sheets created: {', '.join(MARKET_DATA.keys())}")
